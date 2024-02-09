# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl
from openpyxl.utils import get_column_letter
import psycopg2

class RestaurantApp:

    def __init__(self, root):
        
        self.root = root
        self.root.title("Система расчета окупаемости ресторана")

        # Параметры подключения к базе данных
        host = '91.241.13.247'
        port = 1521
        database = 'EDU'
        username = 'intern_team2'
        password = 'fj490#_8gfhfa'

        # Подключение к базе данных
        # URL подключения к базе PostgreSQL
        url = "postgres://bkbppofa:xvLmUmIMfZZ79lLZwN4vFSbpj92SnW6K@manny.db.elephantsql.com/bkbppofa"

        # Подключение к базе данных
        self.conn = psycopg2.connect(url)
        

        # Переменные для данных о пользователе
        self.entry_user_first_name = None
        self.entry_user_last_name = None
        self.entry_user_post = None

        #Переменные для данных о заказчике
        self.client_name = None
        self.client_phone = None
        self.client_email = None

        # Создание переменных для ввода данных о ресторане
        self.entry_name = None
        self.entry_city = None
        self.entry_address = None
        self.entry_avg_check = None
        self.entry_customers_per_day = None

        # Переменные для отслеживания дохода
        self.revenue_per_day = tk.DoubleVar(value=0.00)
        self.revenue_per_month = tk.DoubleVar(value=0.00)

        # Переменные для отслеживания расходов
        self.entry_software_per_month = None
        self.entry_rent_per_month = None
        self.entry_utilities_per_month = None
        self.entry_taxes_per_month = None
        self.entry_salary_per_month = None
        self.entry_purchases_per_month = None
        self.entry_marketing_per_month = None
        self.entry_other_expenses = None
        self.entry_spent_on_launch_total = None

        self.expenses_total = tk.DoubleVar(value=0.00)

        self.active_user_id = None



        # Создание виджетов
        self.create_widgets()

        #self.create_data_display()


    def create_widgets(self):
        # Добавление виджетов для ввода данных о пользователях
        label_user_first_name = ttk.Label(self.root, text="Имя:")
        self.entry_user_first_name = ttk.Entry(self.root)

        label_user_last_name = ttk.Label(self.root, text="Фамилия:")
        self.entry_user_last_name = ttk.Entry(self.root)

        label_user_post = ttk.Label(self.root, text="Должность:")
        self.entry_user_post = ttk.Entry(self.root)

        label_client_separator = ttk.Separator(self.root, orient=tk.HORIZONTAL)
        label_client_separator.grid(row=13, column=0, columnspan=2, pady=10)

        # Добавление виджетов для ввода данных о заказчиках
        label_client_name = ttk.Label(self.root, text="Имя и фамилия:")
        self.entry_client_name = ttk.Entry(self.root)

        label_client_phone = ttk.Label(self.root, text="Телефон:")
        self.entry_client_phone = ttk.Entry(self.root)

        label_client_email = ttk.Label(self.root, text="Email:")
        self.entry_client_email = ttk.Entry(self.root)


        # Создание формы для ввода данных о ресторане
        label_name = ttk.Label(self.root, text="Название ресторана:")
        label_city = ttk.Label(self.root, text="Город:")
        label_address = ttk.Label(self.root, text="Адрес:")

        self.entry_name = ttk.Entry(self.root)
        self.entry_city = ttk.Entry(self.root)
        self.entry_address = ttk.Entry(self.root)

        # поля для таблицы Доходы
        label_avg_check = ttk.Label(self.root, text="Средний чек на человека:")
        self.entry_avg_check = ttk.Entry(self.root)
        label_customers_per_day = ttk.Label(self.root, text="Клиентов в день:")
        self.entry_customers_per_day = ttk.Entry(self.root)

        # Отображение результатов расчета доходов в форме
        self.label_revenue_per_day = ttk.Label(self.root, text="Доход за день:")
        self.label_revenue_per_month = ttk.Label(self.root, text="Доход за месяц:")

        # поля для таблицы Расходы
        label_software_per_month = ttk.Label(self.root, text="Технологии и ПО/мес:")
        self.entry_software_per_month = ttk.Entry(self.root)

        label_rent_per_month = ttk.Label(self.root, text="Аренда/мес:")
        self.entry_rent_per_month = ttk.Entry(self.root)

        label_utilities_per_month = ttk.Label(self.root, text="Коммунальные услуги/мес:")
        self.entry_utilities_per_month = ttk.Entry(self.root)

        label_taxes_per_month = ttk.Label(self.root, text="Налоги/мес:")
        self.entry_taxes_per_month = ttk.Entry(self.root)

        label_salary_per_month = ttk.Label(self.root, text="ФОТ/мес:")
        self.entry_salary_per_month = ttk.Entry(self.root)

        label_purchases_per_month = ttk.Label(self.root, text="Закупка продуктов/мес:")
        self.entry_purchases_per_month = ttk.Entry(self.root)

        label_marketing_per_month = ttk.Label(self.root, text="Маркетинг/мес:")
        self.entry_marketing_per_month = ttk.Entry(self.root)

        label_other_expenses = ttk.Label(self.root, text="Другие расходы:")
        self.entry_other_expenses = ttk.Entry(self.root)

        label_spent_on_launch_total = ttk.Label(self.root, text="Первоначальные инвестиции:")
        self.entry_spent_on_launch_total = ttk.Entry(self.root)

        self.label_expenses_total = ttk.Label(self.root, text="Итого расходы в месяц:")



        # Размещение виджетов на форме
        # Отображение виджетов для ввода данных о пользователях
        # Отображение виджетов для ввода данных о пользователях
        label_users = ttk.Label(self.root, text="Пользователь")
        label_users.grid(row=0, column=0, columnspan=2, pady=10)

        label_user_first_name.grid(row=1, column=0, padx=10, pady=10, sticky=tk.E)
        self.entry_user_first_name.grid(row=1, column=1, padx=10, pady=10, sticky=tk.W)

        label_user_last_name.grid(row=2, column=0, padx=10, pady=10, sticky=tk.E)
        self.entry_user_last_name.grid(row=2, column=1, padx=10, pady=10, sticky=tk.W)

        label_user_post.grid(row=3, column=0, padx=10, pady=10, sticky=tk.E)
        self.entry_user_post.grid(row=3, column=1, padx=10, pady=10, sticky=tk.W)

        # Кнопка "Сохранить пользователя"
        button_save_user = ttk.Button(self.root, text="Сохранить пользователя", command=self.save_user)
        button_save_user.grid(row=4, column=0, columnspan=2, pady=10)

        # Кнопка "Сменить пользователя"
        button_change_user = ttk.Button(self.root, text="Выбрать пользователя", command=self.change_user)
        button_change_user.grid(row=5, column=0, columnspan=2, pady=10)

        # виджеты данных о ресторане
        label_rest = ttk.Label(self.root, text="Ресторан")
        label_rest.grid(row=6, column=0, columnspan=2, pady=10)

        label_name.grid(row=7, column=0, padx=10, pady=10, sticky=tk.E)
        self.entry_name.grid(row=7, column=1, padx=10, pady=10, sticky=tk.W)
        label_city.grid(row=8, column=0, padx=10, pady=10, sticky=tk.E)
        self.entry_city.grid(row=8, column=1, padx=10, pady=10, sticky=tk.W)
        label_address.grid(row=9, column=0, padx=10, pady=10, sticky=tk.E)
        self.entry_address.grid(row=9, column=1, padx=10, pady=10, sticky=tk.W)

        # добавить черту и заголовок Доходы
        label_profit = ttk.Label(self.root, text="Доходы ресторана")
        label_profit.grid(row=10, column=0, columnspan=2, pady=10)

        label_avg_check.grid(row=11, column=0, padx=10, pady=10, sticky=tk.E)
        self.entry_avg_check.grid(row=11, column=1, padx=10, pady=10, sticky=tk.W)
        label_customers_per_day.grid(row=12, column=0, padx=10, pady=10, sticky=tk.E)
        self.entry_customers_per_day.grid(row=12, column=1, padx=10, pady=10, sticky=tk.W)

        self.label_revenue_per_day.grid(row=13, column=0, padx=10, pady=10, sticky=tk.E)
        self.label_revenue_per_month.grid(row=14, column=0, padx=10, pady=10, sticky=tk.E)

        # Отображение разделителя между пользователями и заказчиками
        label_client = ttk.Label(self.root, text="Заказчик")
        label_client.grid(row=0, column=2, columnspan=2, pady=10)

        # Отображение виджетов для ввода данных о заказчиках
        label_client_name.grid(row=1, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_client_name.grid(row=1, column=3, padx=10, pady=10, sticky=tk.W)

        label_client_phone.grid(row=2, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_client_phone.grid(row=2, column=3, padx=10, pady=10, sticky=tk.W)

        label_client_email.grid(row=3, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_client_email.grid(row=3, column=3, padx=10, pady=10, sticky=tk.W)

        # добавить черту и заголовок Расходы
        #line_expenses = ttk.Separator(self.root, orient=tk.HORIZONTAL)
        #line_expenses.grid(row=13, column=0, columnspan=2, pady=10)
        label_expenses = ttk.Label(self.root, text="Расходы ресторана")
        label_expenses.grid(row=4, column=2, columnspan=2, pady=10)

        label_software_per_month.grid(row=5, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_software_per_month.grid(row=5, column=3, padx=10, pady=10, sticky=tk.W)

        label_rent_per_month.grid(row=6, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_rent_per_month.grid(row=6, column=3, padx=10, pady=10, sticky=tk.W)

        label_utilities_per_month.grid(row=7, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_utilities_per_month.grid(row=7, column=3, padx=10, pady=10, sticky=tk.W)

        label_taxes_per_month.grid(row=8, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_taxes_per_month.grid(row=8, column=3, padx=10, pady=10, sticky=tk.W)

        label_salary_per_month.grid(row=9, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_salary_per_month.grid(row=9, column=3, padx=10, pady=10, sticky=tk.W)

        label_purchases_per_month.grid(row=10, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_purchases_per_month.grid(row=10, column=3, padx=10, pady=10, sticky=tk.W)

        label_marketing_per_month.grid(row=11, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_marketing_per_month.grid(row=11, column=3, padx=10, pady=10, sticky=tk.W)

        label_other_expenses.grid(row=12, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_other_expenses.grid(row=12, column=3, padx=10, pady=10, sticky=tk.W)

        label_spent_on_launch_total.grid(row=13, column=2, padx=10, pady=10, sticky=tk.E)
        self.entry_spent_on_launch_total.grid(row=13, column=3, padx=10, pady=10, sticky=tk.W)

        self.label_expenses_total.grid(row=14, column=2, padx=10, pady=10, sticky=tk.W)

        # Кнопка сохранить
        button_save = ttk.Button(self.root, text="Сохранить", command=self.save_data)
        button_save.grid(row=15, column=0, columnspan=1, pady=10)

        # Кнопка отображения результатов рассчета
        button_show_latest_calculation = ttk.Button(self.root, text="Рассчитать показатели",
                                                    command=self.show_latest_calculation)
        button_show_latest_calculation.grid(row=15, column=0, columnspan=2, pady=10)


        button_refresh_data = ttk.Button(self.root, text="Показать последние записи",
                                         command=self.display_data)
        button_refresh_data.grid(row=15, column=1, columnspan=2, pady=10)

        # Добавление Treeview для отображения данных
        columns = (
            "Название ресторана", "Доход за месяц", "Расходы в месяц", "Прибыль", "Первоначальные инвестиции", "Срок окупаемости",
            "Окупаемость в %")
        self.tree = ttk.Treeview(self.root, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor=tk.CENTER)

        self.tree.grid(row=16, column=0, columnspan=4, pady=10)

        # Кнопка сохранить в Excel
        button_save_excel = ttk.Button(self.root, text="Сохранить в Excel", command=self.save_to_excel)
        button_save_excel.grid(row=15, column=2, columnspan=2, pady=10)

    def get_users(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT FirstName || ' ' || LastName FROM Users")
        return [row[0] for row in cursor.fetchall()]

    def change_user(self):
        # Открывает новое окно с выбором пользователя
        self.user_selection_window = tk.Toplevel(self.root)

        # Создание виджетов для выбора пользователя
        label_select_user = ttk.Label(self.user_selection_window, text="Выберите пользователя:")
        label_select_user.grid(row=0, column=0, padx=10, pady=10)

        # Получение списка пользователей из базы данных
        users = self.get_users()

        self.user_var = tk.StringVar(value=users[0] if users else "")
        self.user_combobox = ttk.Combobox(self.user_selection_window, textvariable=self.user_var, values=users)
        self.user_combobox.grid(row=0, column=1, padx=10, pady=10)

        # Кнопка "OK"
        button_ok = ttk.Button(self.user_selection_window, text="OK", command=self.on_user_selected)
        button_ok.grid(row=1, columnspan=2, pady=10)

    def on_user_selected(self):
        # Эта функция вызывается при выборе пользователя в окне выбора пользователя
        selected_user = self.user_var.get()

        if selected_user:
            # Закрытие окна выбора пользователя
            self.user_selection_window.destroy()

            # Установка данных выбранного пользователя в поля ввода
            self.entry_user_first_name.delete(0, tk.END)
            self.entry_user_last_name.delete(0, tk.END)
            self.entry_user_post.delete(0, tk.END)

            first_name, last_name, post = self.get_user_details(selected_user)
            self.entry_user_first_name.insert(0, first_name)
            self.entry_user_last_name.insert(0, last_name)
            self.entry_user_post.insert(0, post)
            # Установка активного пользователя и обновление user_id
            self.set_active_user(selected_user)

    def save_user(self):
        # Сохранение нового пользователя в базе данных
        # Сохранение данных о пользователях
        user_first_name = self.entry_user_first_name.get()
        user_last_name = self.entry_user_last_name.get()
        user_post = self.entry_user_post.get()
        cursor = self.conn.cursor()

        cursor.execute(
            "INSERT INTO Users (FirstName, LastName, Post) VALUES (%s, %s, %s)",
            (user_first_name, user_last_name, user_post)
        )
        self.conn.commit()

    def set_active_user(self, selected_user):
        # Установка активного пользователя
        cursor = self.conn.cursor()
        cursor.execute("SELECT UserID FROM Users WHERE FirstName || ' ' || LastName = %s", (selected_user,))
        user_id = cursor.fetchone()[0]
        self.active_user_id = user_id

    def get_user_details(self, selected_user):
        # Получение данных о выбранном пользователе из базы данных
        cursor = self.conn.cursor()
        cursor.execute("SELECT FirstName, LastName, Post FROM Users WHERE FirstName || ' ' || LastName = %s",
                       (selected_user,))
        user_details = cursor.fetchone()
        return user_details

    def save_data(self):
        cursor = self.conn.cursor()


        # Получение user_id после вставки данных пользователя
        #cursor.execute("SELECT last_insert_rowid()")
        user_id_for_cl = self.active_user_id

        # Сохранение данных о заказчиках
        client_name = self.entry_client_name.get()
        client_phone = self.entry_client_phone.get()
        client_email = self.entry_client_email.get()

        cursor.execute(
            "INSERT INTO Clients (UserID, Name, Phone, Email) VALUES (%s, %s, %s, %s) RETURNING ClientID",
            (user_id_for_cl, client_name, client_phone, client_email)
        )
        client_id = cursor.fetchone()[0]
        self.conn.commit()

        # Получение client_id после вставки данных пользователя
        #cursor.execute("SELECT last_insert_rowid()")
        #client_id = cursor.fetchone()[0]

        # Cохранение данных о ресторане в базу данных
        name = self.entry_name.get()
        city = self.entry_city.get()
        address = self.entry_address.get()

        cursor.execute("INSERT INTO Restaurants (UserID, ClientID, Name_rest, City, Address) "
                       "VALUES (%s, %s, %s, %s, %s) RETURNING RestaurantID", (user_id_for_cl, client_id, name, city, address))
        
        # Получение последнего вставленного ID (RestaurantID) из таблицы Restaurants
        restaurant_id = cursor.fetchone()[0]
        self.conn.commit()


        # Добавление данных о доходах ресторана
        avg_check = float(self.entry_avg_check.get())
        customers_per_day = int(self.entry_customers_per_day.get())

        # Расчет дохода за день
        self.revenue_per_day.set(avg_check * customers_per_day)

        # Обновление отображения дохода за день
        self.label_revenue_per_day.configure(text="Доход за день: {:.2f}".format(self.revenue_per_day.get()))

        # Расчет дохода ресторана за месяц
        self.revenue_per_month.set(avg_check * customers_per_day * 31)
        self.label_revenue_per_month.configure(text="Доход за месяц: {:.2f}".format(self.revenue_per_month.get()))

        # Добавим все данные в таблицу Доходы
        cursor.execute(
            "INSERT INTO Profits (RestaurantID, Avg_Check_Per_Person, "
            "Customers_Per_Day, Revenue_Per_Day, Revenue_Per_Month) VALUES (%s, %s, %s, %s, %s)",
            (restaurant_id, avg_check, customers_per_day, self.revenue_per_day.get(), self.revenue_per_month.get()))
        self.conn.commit()

        # Получение данных о расходах ресторана
        software_per_month = float(self.entry_software_per_month.get())
        rent_per_month = float(self.entry_rent_per_month.get())
        utilities_per_month = float(self.entry_utilities_per_month.get())
        taxes_per_month = float(self.entry_taxes_per_month.get())
        salary_per_month = float(self.entry_salary_per_month.get())
        purchases_per_month = float(self.entry_purchases_per_month.get())
        marketing_per_month = float(self.entry_marketing_per_month.get())
        # Получение значения из виджета "Другие расходы"
        other_expenses_str = self.entry_other_expenses.get()

        # Проверка наличия значения и его преобразование в число
        other_expenses = float(other_expenses_str) if other_expenses_str else 0.0
        spent_on_launch_total = float(self.entry_spent_on_launch_total.get())

        # Рассчет расходов за месяц (Expenses_Total)
        self.expenses_total.set(
                software_per_month +
                rent_per_month +
                utilities_per_month +
                taxes_per_month +
                salary_per_month +
                purchases_per_month +
                marketing_per_month +
                other_expenses
        )

        # Обновление отображения расходов
        self.label_expenses_total.configure(text="Итого расходы в месяц: {:.2f}".format(self.expenses_total.get()))



        # Добавление данных в таблицу Расходы
        cursor.execute(
            '''
            INSERT INTO Expenses (
                RestaurantID, Software_Per_Month, Rent_Per_Month, Utilities_Per_Month,
                Taxes_Per_Month, Salary_Per_Month, Purchases_Per_Month, Marketing_Per_Month,
                Other_Expenses, Spent_On_Launch_Total, Expenses_Total
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''',
            (
                restaurant_id,  # Необходимо определить, откуда брать restaurant_id
                software_per_month,
                rent_per_month,
                utilities_per_month,
                taxes_per_month,
                salary_per_month,
                purchases_per_month,
                marketing_per_month,
                other_expenses,
                spent_on_launch_total,
                self.expenses_total.get()
            )
        )
        # Расчет Payback_Period
        revenue_per_month = self.revenue_per_month.get()
        expenses_total = self.expenses_total.get()
        payback_period = spent_on_launch_total / (revenue_per_month - expenses_total)

        # Расчет ROI
        roi = ((revenue_per_month - expenses_total) * 12) / spent_on_launch_total

        # Добавление данных в таблицу Results
        cursor.execute(
            "INSERT INTO Results (RestaurantID, Payback_Period, ROI) VALUES (%s, %s, %s)",
            (restaurant_id, payback_period, roi)
        )
        self.conn.commit()

    def show_latest_calculation(self):
        # Получение последней записи из таблицы Restaurants
        cursor = self.conn.cursor()
        cursor.execute(
            """
            SELECT Restaurants.Name_rest, Profits.Revenue_Per_Month, Expenses.Expenses_Total,
                (Profits.Revenue_Per_Month-Expenses.Expenses_Total) as profit,
                Expenses.Spent_On_Launch_Total, ROUND(Results.Payback_Period::numeric, 2),
                ROUND(Results.ROI::numeric * 100, 2) AS ROI_percent
            FROM Restaurants
            LEFT JOIN Profits ON Restaurants.RestaurantID = Profits.RestaurantID
            LEFT JOIN Expenses ON Restaurants.RestaurantID = Expenses.RestaurantID
            LEFT JOIN Results ON Restaurants.RestaurantID = Results.RestaurantID
            ORDER BY Restaurants.RestaurantID DESC
            LIMIT 1;
            """
        )

        row = cursor.fetchone()

        # Очищаем Treeview перед обновлением
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Вставляем данные в Treeview
        if row:
            self.tree.insert("", tk.END, values=row)

    def display_data(self):
        # Очищаем Treeview перед обновлением
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Получаем данные из базы
        cursor = self.conn.cursor()
        cursor.execute(
                """
                SELECT Restaurants.Name_rest, Profits.Revenue_Per_Month, Expenses.Expenses_Total,
                       (Profits.Revenue_Per_Month-Expenses.Expenses_Total) as profit, 
                       Expenses.Spent_On_Launch_Total, ROUND(Results.Payback_Period::numeric, 2),
                       ROUND(Results.ROI::numeric * 100, 2) AS ROI_percent
                FROM Restaurants
                LEFT JOIN Profits ON Restaurants.RestaurantID = Profits.RestaurantID
                LEFT JOIN Expenses ON Restaurants.RestaurantID = Expenses.RestaurantID
                LEFT JOIN Results ON Restaurants.RestaurantID = Results.RestaurantID
                """
            )

        rows = cursor.fetchall()

        # Вставляем данные в Treeview
        for row in rows:
            self.tree.insert("", tk.END, values=row)

        # Очистка полей ввода
        #self.entry_name.delete(0, tk.END)
        #self.entry_city.delete(0, tk.END)
        #self.entry_address.delete(0, tk.END)
        #self.entry_avg_check.delete(0, tk.END)
        #self.entry_customers_per_day.delete(0, tk.END)
            
    def save_to_excel(self):
        # Запрос для получения данных
        query = """
        SELECT Restaurants.Name_rest, Profits.Revenue_Per_Month, Expenses.Expenses_Total,
            (Profits.Revenue_Per_Month-Expenses.Expenses_Total) as profit,
            Expenses.Spent_On_Launch_Total, ROUND(Results.Payback_Period::numeric, 2),
            ROUND(Results.ROI::numeric * 100, 2) AS ROI_percent
        FROM Restaurants
        LEFT JOIN Profits ON Restaurants.RestaurantID = Profits.RestaurantID
        LEFT JOIN Expenses ON Restaurants.RestaurantID = Expenses.RestaurantID
        LEFT JOIN Results ON Restaurants.RestaurantID = Results.RestaurantID
        ORDER BY Restaurants.RestaurantID DESC
        LIMIT 1;
        """

        cursor = self.conn.cursor()
        cursor.execute(query)
        data = cursor.fetchone()

        # Выбор пути сохранения файла
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        if file_path:
            # Создание нового Excel-файла
            wb = openpyxl.Workbook()
            ws = wb.active

            # Заголовки
            headers = ["Название ресторана", "Доход за месяц", "Расходы в месяц", "Прибыль", "Первоначальные инвестиции", "Срок окупаемости",
            "Окупаемость в %"]
            ws.append(headers)

            # Добавление данных
            ws.append(data)

            # Установка ширины столбцов
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                max_length = max(len(str(header)), len(str(data[col_num - 1])))  # Находим максимальную длину
                adjusted_width = (max_length + 2) * 1.2  # Настройте коэффициент по вашему усмотрению
                ws.column_dimensions[col_letter].width = adjusted_width


            # Сохранение файла
            wb.save(file_path)

            print("Данные сохранены в Excel.")
    
    def run(self):
        # Запуск главного цикла
        self.root.mainloop()

    def __del__(self):
        # Закрытие соединения при удалении объекта
        self.conn.close()

# Создание приложения и запуск
root = tk.Tk()
app = RestaurantApp(root)
app.run()
    



